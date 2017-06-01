
import re
import datetime
import csv

from openpyxl.workbook import Workbook
from openpyxl.styles import numbers, is_date_format, Style
from openpyxl.utils.datetime import to_excel

#import gen_functions

class Linewriter:

    def __init__(self, lines):
        self.lines = lines
        self.current_workbook = False

    def set_lines(self, lines):
        self.lines = lines

    def write_xlsx(self, headers, header_style, outfile, encoding='utf-8'):
        if not self.current_workbook:
            self.current_workbook = Workbook(encoding = encoding)
            ws = self.current_workbook.active
            ws.title = 'sheet1'
        else:
            num_sheets = len(self.current_workbook.get_sheet_names())
            ws = self.current_workbook.create_sheet(title='sheet'+str(num_sheets+1))
        number_header = {}
        for i, header in enumerate(headers):
            i += 1
            number_header[i] = header
            _cell = ws.cell(column = i, row = 1, value = header)
        for i, line in enumerate(self.lines):
#            print('\t'.join([str(x) for x in line]).encode('utf-8'))
            i += 2
            for j, col in enumerate(line):
                # try:
                #     print(col)
                # except:
                #     print(col.encode('utf-8'))
                j += 1
#                print(type(col))
#                if j != 1:
#                    print(i, j, col)
                try:
                    _cell = ws.cell(row = i, column = j, value = col)
                except:
                    _cell = ws.cell(row=i,column=j,value='')
                    continue
                if re.search('^http', str(col)) and not ' ' in str(col):
                    #_cell.hyperlink = col
                    _cell = ws.cell(row = i, column = j, value = '=HYPERLINK("' + col + '","' + col + '")')
                else:
                    _cell = ws.cell(row = i, column = j, value = col)
                    st = header_style[number_header[j]]
 #                   print(col, st)
                    if not st == 'general':
                        if ':' in st:
                            _cell.number_format = numbers.FORMAT_DATE_TIME6
                        elif '-' in st:
                            _cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
                        else:
                            #try:
                            #    _cell = ws.cell(row = i, column = j, value = int(col))
                            #except:
                            #    _cell = ws.cell(row = i, column = j, value = float(col))
                            _cell.number_format = st
        self.current_workbook.save(filename = outfile)

    def write_txt(self, outfile, delimiter = '\t'):
        with open(outfile, 'w', encoding = 'utf-8') as out:
            for line in self.lines:
                out.write(delimiter.join([str(x) for x in line]) + '\n')

    def write_csv(self, outfile):
        """
        CSV writer
        =====
        Function to write rows to a file in csv format

        Parameters
        -----
        rows : list of lists
        outfile : str
            The name of the file to write the rows to
        """
        with open(outfile, 'w', encoding = 'utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for line in self.lines:
                writer.writerow(line)
